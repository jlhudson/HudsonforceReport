import os
import platform
import re
import subprocess
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from Levenshtein import ratio
from openpyxl.styles import Font, PatternFill

from app.reportlogger import report_logger
from app.reports.shift_change_tracker import ShiftChangeTracker


class UnfilledShiftReport:
    PRIORITIES = {'critical': {'max_days': 2, 'color': 'FF0000'},  # Red
                  'urgent': {'max_days': 5, 'color': '000000'},  # Black
                  'standard': {'max_days': 8, 'color': '000000'}  # Black
                  }

    def __init__(self, dataset):
        self.dataset = dataset
        self.report_path = Path("Humanforce Reports")
        self.ensure_report_directory()
        self.working_days_ahead = 8
        self.shift_tracker = ShiftChangeTracker()

    def ensure_report_directory(self):
        """Ensure the report directory exists."""
        self.report_path.mkdir(parents=True, exist_ok=True)

    def get_priority(self, shift_date):
        """Determine priority based on days ahead."""
        today = datetime.now().date()
        days_until = (shift_date - today).days

        if days_until <= 2:
            return 'critical'
        elif days_until <= 5:
            return 'urgent'
        else:
            return 'standard'

    def get_escalation_status(self, comment):
        if not comment or pd.isna(comment) or not isinstance(comment, str):
            return "Not Escalated"

        comment = str(comment).lower()
        lines = comment.split('\n')[:5]

        # Regex patterns
        escalated_patterns = [r'\besc\b', r'\bescalated\b', r'\besclated\b', r'\beskalated\b']
        ready_patterns = [r'ready.*(?:esc|escalate)', r'(?:ready|rdy).*(?:2|to).*(?:esc|escalate)']

        # Try regex first
        for line in lines:
            if any(re.search(pattern, line, re.I) for pattern in escalated_patterns):
                return "Escalated"
            if any(re.search(pattern, line, re.I) for pattern in ready_patterns):
                return "Ready to Escalate"

        # If no regex match, try Levenshtein
        words = set(' '.join(lines).split())
        escalated_words = {'esc', 'escalated', 'escalation', 'esclated'}
        ready_words = {'ready2escalate', 'readytoescalate', 'rdy2esc'}

        for word in words:
            if any(ratio(word, esc_word) > 0.8 for esc_word in escalated_words):
                return "Escalated"
            if any(ratio(word, ready_word) > 0.8 for ready_word in ready_words):
                return "Ready to Escalate"

        return "Not Escalated"

    def format_date(self, date):
        """Format date in South Australian standard."""
        return date.strftime('%a, %d/%m')

    def open_file(self, file_path):
        """Open the generated Excel file using the default system application."""
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', file_path])
            elif platform.system() == 'Windows':  # Windows
                os.startfile(file_path)
            else:  # Linux variants
                subprocess.run(['xdg-open', file_path])
            report_logger.info(f"Opened file: {file_path}")
        except Exception as e:
            report_logger.error(f"Error opening file: {e}")

    def generate_report(self):
        """Generate the unfilled shifts report in Excel format."""
        self.shift_tracker.process_shifts(self.dataset.unassigned_shifts)

        today = datetime.now().date()
        end_date = today + timedelta(days=14)
        shifts_by_location = defaultdict(list)

        for shift in self.dataset.unassigned_shifts:
            shift_date = shift.start.date()
            if today <= shift_date <= end_date:
                priority = self.get_priority(shift_date)
                shifts_by_location[shift.work_area.location].append(
                    {'Department': shift.work_area.department,
                     'Start Date': shift_date,
                     'Start': shift.start.strftime('%H:%M'),
                     'End': shift.end.strftime('%H:%M'),
                     'Role': shift.work_area.role,
                     'Priority': priority.title(),
                     'Color': self.PRIORITIES[priority]['color'],
                     'Status': self.get_escalation_status(shift.comment)})

        if not shifts_by_location:
            report_logger.info("No unfilled shifts found for the next 14 days.")
            return None

        current_time = datetime.now().strftime('%d %b')
        excel_path = self.report_path / f'Unfilled Shifts Report {current_time}.xlsx'

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for location, shifts in shifts_by_location.items():
                if not shifts:
                    continue

                df = pd.DataFrame(shifts)
                df['Start Date'] = pd.to_datetime(df['Start Date'])
                df['Start DateTime'] = df.apply(
                    lambda x: x['Start Date'].replace(hour=int(x['Start'].split(':')[0]),
                                                      minute=int(x['Start'].split(':')[1])), axis=1)

                dept_avg_times = df.groupby('Department')['Start DateTime'].mean().reset_index()
                dept_avg_times = dept_avg_times.sort_values('Start DateTime')
                dept_order = {dept: idx for idx, dept in enumerate(dept_avg_times['Department'])}
                df['Dept_Rank'] = df['Department'].map(dept_order)

                df = df.sort_values(['Dept_Rank', 'Start Date', 'Start'])
                df = df.drop(['Start DateTime', 'Dept_Rank'], axis=1)
                df['Start Date'] = df['Start Date'].apply(self.format_date)

                df_excel = df.drop(['Color'], axis=1)
                column_order = ['Priority', 'Department', 'Role', 'Start Date', 'Start', 'End', 'Status']
                df_excel = df_excel[column_order]

                sheet_name = f"{str(location)[:25]} ({len(shifts)})"[:31]
                df_excel.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                self._format_worksheet(worksheet, df)

        report_logger.info(f"\nExcel report generated: {excel_path}")
        self.open_file(excel_path)  # Open the file after creation
        return excel_path

    def _format_worksheet(self, worksheet, df):
        """Apply formatting to the worksheet with consistent row coloring."""
        # Remove all borders from the worksheet
        from openpyxl.styles import Border, Side
        no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

        # Format header row with bold black text and no borders
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='000000')
            cell.border = no_border

        # Format data rows based on priority
        for excel_row, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            priority_cell_value = row[0].value

            row_color = '000000'
            if priority_cell_value == 'Critical':
                row_color = 'FF0000'

            for cell in row:
                cell.font = Font(color=row_color, bold=False)
                cell.fill = PatternFill(fill_type=None)
                cell.border = no_border

        # Auto-size columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
